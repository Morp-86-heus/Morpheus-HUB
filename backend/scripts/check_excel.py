import openpyxl
wb = openpyxl.load_workbook('/app/data/Interventi_Tecnici.xlsm', read_only=True, keep_vba=True)
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    print(f'--- {name} ---')
    for r in rows:
        print(r)
    print()
