from datetime import datetime
from io import BytesIO
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Response, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_
from database import get_db
from models import Articolo, MovimentoMagazzino, Organizzazione, SottoMagazzino, User, RuoloEnum
from auth import require_roles, get_active_org_id
from utils.plan import check_feature
import schemas
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter(prefix="/api/magazzino", tags=["magazzino"])

_tutti = require_roles(RuoloEnum.proprietario, RuoloEnum.amministratore, RuoloEnum.commerciale, RuoloEnum.tecnico)
_admin = require_roles(RuoloEnum.proprietario, RuoloEnum.amministratore)


@router.get("/articoli", response_model=List[schemas.ArticoloOut])
def list_articoli(
    commitente: Optional[str] = None,
    cliente: Optional[str] = None,
    categoria: Optional[str] = None,
    search: Optional[str] = None,
    sotto_minimo: Optional[bool] = None,
    db: Session = Depends(get_db),
    _: User = Depends(_tutti),
    org_id: int = Depends(get_active_org_id),
):
    q = db.query(Articolo).options(
        joinedload(Articolo.movimenti),
        joinedload(Articolo.sotto_magazzino),
    ).filter(Articolo.organizzazione_id == org_id)
    if commitente:
        q = q.filter(Articolo.commitente == commitente)
    if cliente:
        q = q.filter(Articolo.cliente == cliente)
    if categoria:
        q = q.filter(Articolo.categoria == categoria)
    if search:
        pattern = f"%{search}%"
        q = q.filter(or_(
            Articolo.seriale.ilike(pattern),
            Articolo.cespite.ilike(pattern),
            Articolo.marca.ilike(pattern),
            Articolo.modello.ilike(pattern),
            Articolo.descrizione.ilike(pattern),
            Articolo.fornitore.ilike(pattern),
        ))
    if sotto_minimo is True:
        q = q.filter(
            Articolo.quantita_minima.isnot(None),
            Articolo.quantita_disponibile <= Articolo.quantita_minima,
        )
    return q.order_by(Articolo.commitente, Articolo.cliente, Articolo.categoria, Articolo.descrizione).all()


@router.post("/articoli", response_model=schemas.ArticoloOut, status_code=201)
def create_articolo(
    payload: schemas.ArticoloCreate,
    db: Session = Depends(get_db),
    _: User = Depends(_admin),
    org_id: int = Depends(get_active_org_id),
):
    org = db.query(Organizzazione).filter_by(id=org_id).first()
    check_feature(org, 'magazzino')
    if payload.seriale:
        if db.query(Articolo).filter(
            Articolo.organizzazione_id == org_id,
            Articolo.seriale == payload.seriale,
        ).first():
            raise HTTPException(status_code=409, detail=f"Seriale '{payload.seriale}' già presente in magazzino")
    if payload.cespite:
        if db.query(Articolo).filter(
            Articolo.organizzazione_id == org_id,
            Articolo.cespite == payload.cespite,
        ).first():
            raise HTTPException(status_code=409, detail=f"Cespite '{payload.cespite}' già presente in magazzino")
    obj = Articolo(**payload.model_dump(), organizzazione_id=org_id)
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return db.query(Articolo).options(joinedload(Articolo.movimenti), joinedload(Articolo.sotto_magazzino)).filter_by(id=obj.id).first()


@router.get("/articoli/template")
def download_template(
    _: User = Depends(_admin),
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Articoli"

    headers = [
        ("Commitente*", True),
        ("Cliente*", True),
        ("Categoria", False),
        ("Marca", False),
        ("Modello", False),
        ("Seriale", False),
        ("Cespite", False),
        ("Descrizione*", True),
        ("Unita_Misura", False),
        ("Quantita_Iniziale", False),
        ("Quantita_Minima", False),
        ("Fornitore", False),
        ("Note", False),
    ]

    for col, (name, required) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="1D4ED8" if required else "6B7280")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    example = ["BPP", "SOGEI", "Stampanti", "HP", "LaserJet 400",
               "SN123456", "", "Toner HP LaserJet 400", "pz", 5, 2, "HP Italia", ""]
    ws.append(example)
    for col in range(1, len(headers) + 1):
        ws.cell(row=2, column=col).font = Font(color="6B7280", italic=True)

    col_widths = [16, 16, 14, 12, 16, 16, 14, 32, 13, 17, 15, 16, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_carico_massivo.xlsx"},
    )


IMPORT_HEADER_MAP = {
    "commitente*": "commitente", "commitente": "commitente",
    "cliente*": "cliente", "cliente": "cliente",
    "categoria": "categoria",
    "marca": "marca",
    "modello": "modello",
    "seriale": "seriale",
    "cespite": "cespite",
    "descrizione*": "descrizione", "descrizione": "descrizione",
    "unita_misura": "unita_misura",
    "quantita_iniziale": "quantita_disponibile",
    "quantita_minima": "quantita_minima",
    "fornitore": "fornitore",
    "note": "note",
}


def _parse_import_file(content: bytes, org_id: int, db: Session):
    try:
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "File non valido. Carica un file .xlsx")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "Il file non contiene dati (manca la riga intestazione o le righe dati)")

    raw_headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    field_names = [IMPORT_HEADER_MAP.get(h) for h in raw_headers]

    # Precarica seriali e cespiti esistenti per controllo rapido
    existing_seriali = {
        r[0] for r in db.query(Articolo.seriale)
        .filter(Articolo.organizzazione_id == org_id, Articolo.seriale.isnot(None)).all()
    }
    existing_cespiti = {
        r[0] for r in db.query(Articolo.cespite)
        .filter(Articolo.organizzazione_id == org_id, Articolo.cespite.isnot(None)).all()
    }

    # Traccia duplicati all'interno dello stesso file
    file_seriali: set = set()
    file_cespiti: set = set()

    preview = []
    for row_idx, row in enumerate(rows[1:], start=2):
        if all(v is None or str(v).strip() == "" for v in row):
            continue  # salta righe vuote

        data: dict = {}
        for fi, field in enumerate(field_names):
            if field and fi < len(row) and row[fi] is not None:
                val = row[fi]
                data[field] = str(val).strip() if not isinstance(val, (int, float)) else val

        def _int(k):
            try:
                return int(data.get(k) or 0)
            except (ValueError, TypeError):
                return 0

        errore = None
        if not data.get("commitente"):
            errore = "Commitente obbligatorio"
        elif not data.get("cliente"):
            errore = "Cliente obbligatorio"
        elif not data.get("descrizione"):
            errore = "Descrizione obbligatoria"
        elif data.get("seriale") and data["seriale"] in existing_seriali:
            errore = f"Seriale '{data['seriale']}' già presente in magazzino"
        elif data.get("seriale") and data["seriale"] in file_seriali:
            errore = f"Seriale '{data['seriale']}' duplicato nel file"
        elif data.get("cespite") and data["cespite"] in existing_cespiti:
            errore = f"Cespite '{data['cespite']}' già presente in magazzino"
        elif data.get("cespite") and data["cespite"] in file_cespiti:
            errore = f"Cespite '{data['cespite']}' duplicato nel file"

        if errore is None:
            if data.get("seriale"):
                file_seriali.add(data["seriale"])
            if data.get("cespite"):
                file_cespiti.add(data["cespite"])

        preview.append({
            "riga": row_idx,
            "valido": errore is None,
            "errore": errore,
            "commitente": data.get("commitente", ""),
            "cliente": data.get("cliente", ""),
            "categoria": data.get("categoria", ""),
            "marca": data.get("marca", ""),
            "modello": data.get("modello", ""),
            "seriale": data.get("seriale", ""),
            "cespite": data.get("cespite", ""),
            "descrizione": data.get("descrizione", ""),
            "unita_misura": data.get("unita_misura") or "pz",
            "quantita_disponibile": _int("quantita_disponibile"),
            "quantita_minima": _int("quantita_minima"),
            "fornitore": data.get("fornitore", ""),
            "note": data.get("note", ""),
        })

    return preview


@router.post("/articoli/import")
async def import_articoli(
    file: UploadFile = File(...),
    dry_run: bool = False,
    db: Session = Depends(get_db),
    _: User = Depends(_admin),
    org_id: int = Depends(get_active_org_id),
):
    org = db.query(Organizzazione).filter_by(id=org_id).first()
    check_feature(org, "magazzino")
    content = await file.read()
    preview = _parse_import_file(content, org_id, db)

    validi = sum(1 for r in preview if r["valido"])
    invalidi = len(preview) - validi

    if dry_run:
        return {"righe": preview, "validi": validi, "invalidi": invalidi}

    importati = 0
    saltati = []
    for r in preview:
        if not r["valido"]:
            saltati.append({"riga": r["riga"], "motivo": r["errore"]})
            continue
        db.add(Articolo(
            commitente=r["commitente"],
            cliente=r["cliente"],
            categoria=r["categoria"] or None,
            marca=r["marca"] or None,
            modello=r["modello"] or None,
            seriale=r["seriale"] or None,
            cespite=r["cespite"] or None,
            descrizione=r["descrizione"],
            unita_misura=r["unita_misura"],
            quantita_disponibile=r["quantita_disponibile"],
            quantita_minima=r["quantita_minima"],
            fornitore=r["fornitore"] or None,
            note=r["note"] or None,
            organizzazione_id=org_id,
        ))
        importati += 1
    db.commit()
    return {"importati": importati, "saltati": saltati}


@router.put("/articoli/{aid}", response_model=schemas.ArticoloOut)
def update_articolo(
    aid: int,
    payload: schemas.ArticoloUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(_admin),
    org_id: int = Depends(get_active_org_id),
):
    org = db.query(Organizzazione).filter_by(id=org_id).first()
    check_feature(org, 'magazzino')
    obj = db.query(Articolo).filter_by(id=aid, organizzazione_id=org_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Articolo non trovato")
    data = payload.model_dump(exclude_unset=True)
    if data.get("seriale") and data["seriale"] != obj.seriale:
        if db.query(Articolo).filter(
            Articolo.organizzazione_id == org_id,
            Articolo.seriale == data["seriale"],
            Articolo.id != aid,
        ).first():
            raise HTTPException(status_code=409, detail=f"Seriale '{data['seriale']}' già presente in magazzino")
    if data.get("cespite") and data["cespite"] != obj.cespite:
        if db.query(Articolo).filter(
            Articolo.organizzazione_id == org_id,
            Articolo.cespite == data["cespite"],
            Articolo.id != aid,
        ).first():
            raise HTTPException(status_code=409, detail=f"Cespite '{data['cespite']}' già presente in magazzino")
    for k, v in data.items():
        setattr(obj, k, v)
    obj.updated_at = datetime.utcnow()
    db.commit()
    return db.query(Articolo).options(joinedload(Articolo.movimenti), joinedload(Articolo.sotto_magazzino)).filter_by(id=aid).first()


@router.delete("/articoli/{aid}", status_code=204)
def delete_articolo(
    aid: int,
    db: Session = Depends(get_db),
    _: User = Depends(_admin),
    org_id: int = Depends(get_active_org_id),
):
    org = db.query(Organizzazione).filter_by(id=org_id).first()
    check_feature(org, 'magazzino')
    obj = db.query(Articolo).filter_by(id=aid, organizzazione_id=org_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Articolo non trovato")
    db.delete(obj)
    db.commit()


@router.post("/articoli/{aid}/movimenti", status_code=201)
def add_movimento(
    aid: int,
    payload: schemas.MovimentoCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(_tutti),
    org_id: int = Depends(get_active_org_id),
):
    org = db.query(Organizzazione).filter_by(id=org_id).first()
    check_feature(org, 'magazzino')
    articolo = db.query(Articolo).filter_by(id=aid, organizzazione_id=org_id).first()
    if not articolo:
        raise HTTPException(status_code=404, detail="Articolo non trovato")

    if payload.tipo not in ("carico", "scarico", "rettifica"):
        raise HTTPException(status_code=400, detail="Tipo movimento non valido. Usa: carico, scarico, rettifica")

    if payload.tipo == "scarico" and articolo.quantita_disponibile < payload.quantita:
        raise HTTPException(status_code=400, detail="Quantità in scarico superiore alla giacenza disponibile")

    is_serialized = bool(articolo.seriale or articolo.cespite)

    db.add(MovimentoMagazzino(
        articolo_id=aid,
        tipo=payload.tipo,
        quantita=payload.quantita,
        riferimento_ticket_id=payload.riferimento_ticket_id,
        note=payload.note,
        creato_da=current_user.id,
    ))

    if payload.tipo == "carico":
        articolo.quantita_disponibile += payload.quantita
    elif payload.tipo == "scarico":
        articolo.quantita_disponibile -= payload.quantita
    else:
        articolo.quantita_disponibile = payload.quantita

    if is_serialized and articolo.quantita_disponibile <= 0 and payload.tipo in ("scarico", "rettifica"):
        db.commit()
        db.delete(articolo)
        db.commit()
        return Response(status_code=204)

    articolo.updated_at = datetime.utcnow()
    db.commit()
    return db.query(Articolo).options(joinedload(Articolo.movimenti), joinedload(Articolo.sotto_magazzino)).filter_by(id=aid).first()


@router.get("/cerca-articolo")
def cerca_articolo(
    seriale: Optional[str] = None,
    cespite: Optional[str] = None,
    commitente: Optional[str] = None,
    cliente: Optional[str] = None,
    db: Session = Depends(get_db),
    _: User = Depends(_tutti),
    org_id: int = Depends(get_active_org_id),
):
    if not seriale and not cespite:
        return []
    q = db.query(Articolo).filter(Articolo.organizzazione_id == org_id)
    if seriale:
        q = q.filter(Articolo.seriale == seriale)
    elif cespite:
        q = q.filter(Articolo.cespite == cespite)
    if commitente:
        q = q.filter(Articolo.commitente == commitente)
    if cliente:
        q = q.filter(Articolo.cliente == cliente)
    q = q.filter(or_(Articolo.categoria != "Gestione Guasti", Articolo.categoria.is_(None)))
    return q.limit(5).all()


@router.get("/movimenti", response_model=schemas.PaginatedMovimenti)
def list_movimenti(
    commitente: Optional[str] = None,
    cliente: Optional[str] = None,
    tipo: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    db: Session = Depends(get_db),
    _: User = Depends(_admin),
    org_id: int = Depends(get_active_org_id),
):
    q = (db.query(MovimentoMagazzino)
           .join(Articolo, MovimentoMagazzino.articolo_id == Articolo.id)
           .outerjoin(User, MovimentoMagazzino.creato_da == User.id)
           .filter(Articolo.organizzazione_id == org_id))
    if commitente:
        q = q.filter(Articolo.commitente == commitente)
    if cliente:
        q = q.filter(Articolo.cliente == cliente)
    if tipo:
        q = q.filter(MovimentoMagazzino.tipo == tipo)
    if search:
        pattern = f"%{search}%"
        q = q.filter(or_(
            Articolo.descrizione.ilike(pattern),
            Articolo.seriale.ilike(pattern),
            Articolo.modello.ilike(pattern),
            MovimentoMagazzino.note.ilike(pattern),
        ))
    total = q.count()
    rows = q.order_by(MovimentoMagazzino.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    items = []
    for m in rows:
        utente = db.get(User, m.creato_da) if m.creato_da else None
        items.append(schemas.MovimentoLogOut(
            id=m.id,
            tipo=m.tipo,
            quantita=m.quantita,
            note=m.note,
            riferimento_ticket_id=m.riferimento_ticket_id,
            created_at=m.created_at,
            articolo_id=m.articolo_id,
            articolo_descrizione=m.articolo.descrizione,
            articolo_marca=m.articolo.marca,
            articolo_modello=m.articolo.modello,
            articolo_seriale=m.articolo.seriale,
            articolo_categoria=m.articolo.categoria,
            articolo_commitente=m.articolo.commitente,
            articolo_cliente=m.articolo.cliente,
            creato_da_nome=utente.nome_completo if utente else None,
        ))
    return {"total": total, "page": page, "page_size": page_size, "items": items}


@router.get("/categorie")
def list_categorie(
    db: Session = Depends(get_db),
    _: User = Depends(_tutti),
    org_id: int = Depends(get_active_org_id),
):
    rows = (
        db.query(Articolo.categoria)
        .filter(Articolo.organizzazione_id == org_id, Articolo.categoria.isnot(None))
        .distinct()
        .order_by(Articolo.categoria)
        .all()
    )
    return [r[0] for r in rows]


# ── Ubicazioni ─────────────────────────────────────────────────────────────────

@router.get("/sotto-magazzini", response_model=List[schemas.SottoMagazzinoOut])
def list_sotto_magazzini(
    commitente: Optional[str] = None,
    db: Session = Depends(get_db),
    _: User = Depends(_tutti),
    org_id: int = Depends(get_active_org_id),
):
    q = db.query(SottoMagazzino).filter(SottoMagazzino.organizzazione_id == org_id)
    if commitente:
        q = q.filter(SottoMagazzino.commitente == commitente)
    return q.order_by(SottoMagazzino.commitente, SottoMagazzino.nome).all()


@router.post("/sotto-magazzini", response_model=schemas.SottoMagazzinoOut, status_code=201)
def create_sotto_magazzino(
    payload: schemas.SottoMagazzinoCreate,
    db: Session = Depends(get_db),
    _: User = Depends(_admin),
    org_id: int = Depends(get_active_org_id),
):
    org = db.query(Organizzazione).filter_by(id=org_id).first()
    check_feature(org, 'magazzino')
    obj = SottoMagazzino(**payload.model_dump(), organizzazione_id=org_id)
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@router.put("/sotto-magazzini/{sid}", response_model=schemas.SottoMagazzinoOut)
def update_sotto_magazzino(
    sid: int,
    payload: schemas.SottoMagazzinoUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(_admin),
    org_id: int = Depends(get_active_org_id),
):
    obj = db.query(SottoMagazzino).filter_by(id=sid, organizzazione_id=org_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Ubicazione non trovata")
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(obj, k, v)
    obj.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(obj)
    return obj


@router.delete("/sotto-magazzini/{sid}", status_code=204)
def delete_sotto_magazzino(
    sid: int,
    db: Session = Depends(get_db),
    _: User = Depends(_admin),
    org_id: int = Depends(get_active_org_id),
):
    obj = db.query(SottoMagazzino).filter_by(id=sid, organizzazione_id=org_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Ubicazione non trovata")
    # Gli articoli che puntano a questa ubicazione tornano al principale (SET NULL via FK)
    db.delete(obj)
    db.commit()


# ── Spostamento articolo ───────────────────────────────────────────────────────

@router.post("/articoli/{aid}/sposta", response_model=schemas.ArticoloOut)
def sposta_articolo(
    aid: int,
    payload: schemas.SpostamentoCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(_admin),
    org_id: int = Depends(get_active_org_id),
):
    articolo = db.query(Articolo).filter_by(id=aid, organizzazione_id=org_id).first()
    if not articolo:
        raise HTTPException(status_code=404, detail="Articolo non trovato")

    dest = None
    if payload.sotto_magazzino_id is not None:
        dest = db.query(SottoMagazzino).filter_by(id=payload.sotto_magazzino_id, organizzazione_id=org_id).first()
        if not dest:
            raise HTTPException(status_code=404, detail="Ubicazione non trovata")
        if dest.commitente != articolo.commitente:
            raise HTTPException(status_code=400, detail="L'ubicazione non appartiene allo stesso committente dell'articolo")

    provenienza = articolo.sotto_magazzino.nome if articolo.sotto_magazzino else "Magazzino principale"
    destinazione = dest.nome if dest else "Magazzino principale"
    nota_auto = f"Spostato da '{provenienza}' a '{destinazione}'"
    if payload.note:
        nota_auto += f" — {payload.note}"

    db.add(MovimentoMagazzino(
        articolo_id=aid,
        tipo="trasferimento",
        quantita=articolo.quantita_disponibile,
        note=nota_auto,
        creato_da=current_user.id,
    ))

    articolo.sotto_magazzino_id = payload.sotto_magazzino_id
    articolo.updated_at = datetime.utcnow()
    db.commit()
    return db.query(Articolo).options(joinedload(Articolo.movimenti), joinedload(Articolo.sotto_magazzino)).filter_by(id=aid).first()
